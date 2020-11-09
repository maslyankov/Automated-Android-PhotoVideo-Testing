import os

from openpyxl import Workbook, load_workbook
from openpyxl import cell as xlcell, worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, PatternFill, Font, Border, Alignment, Side
from openpyxl.drawing.image import Image as xls_image

import win32com.client as win32

# Local
import src.constants as constants
from src.app.utils import kelvin_to_illumenant, only_digits, only_chars, extract_video_frame


def xls_to_xlsx(xls_file) -> str:
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(xls_file)
    new_file = xls_file + "x"

    wb.SaveAs(new_file, FileFormat=51)  # FileFormat = 51 is for .xlsx extension
    wb.Close()  # FileFormat = 56 is for .xls extension
    excel.Application.Quit()

    return new_file


def within_range(bounds: tuple, cell: xlcell) -> bool:
    column_start, row_start, column_end, row_end = bounds
    row = cell.row
    if row_start <= row <= row_end:
        column = cell.column
        if column_start <= column <= column_end:
            return True
    return False


def get_value_merged(sheet: worksheet, cell: xlcell) -> any:
    for merged in sheet.merged_cells:
        if within_range(merged.bounds, cell):
            return sheet.cell(merged.min_row, merged.min_col).value
    return cell.value


def is_merged(cell) -> bool:
    if type(cell).__name__ == 'MergedCell':
        return True
    else:
        return False


def get_cell_val(ws, cell):
    return get_value_merged(ws, cell) if is_merged(cell) else cell.value


def parse_excel_template(excel_file) -> dict:
    # Config
    header_mapping = {
        'test_type': "test target",
        'light_temp': "light",
        'lux': "lux",
        'param': "param",
        'min': "min",
        'max': "max"
    }
    conf_min_col = 2
    conf_min_row = 2
    conf_max_row = 77
    filter_str = 'functional'
    # Config end

    excel_file = os.path.normpath(excel_file)
    print(f'Parsing "{excel_file}"...')

    ext = excel_file.split('.')[-1]

    if ext == 'xls':
        print("Got .XLS.. Converting it to XLSX")
        excel_file = xls_to_xlsx(excel_file)

    wb = load_workbook(filename=excel_file)
    ws = wb.worksheets[0]

    tests_seq = {}

    header = list(ws.rows)[conf_min_row - 1]
    print(list(ws.rows))

    test_type_col = None
    light_temp_col = None
    lux_col = None
    param_col = None
    min_col = None
    max_col = None

    for cell in header:
        if cell.column < conf_min_col:
            continue
        print('value is: ', cell.value)
        if header_mapping['test_type'] in cell.value.lower():
            test_type_col = cell.column_letter

        if header_mapping['light_temp'] in cell.value.lower():
            light_temp_col = cell.column_letter

        if header_mapping['lux'] in cell.value.lower():
            lux_col = cell.column_letter

        if header_mapping['param'] in cell.value.lower():
            param_col = cell.column_letter

        if header_mapping['min'] in cell.value.lower():
            min_col = cell.column_letter

        if header_mapping['max'] in cell.value.lower():
            max_col = cell.column_letter

    current_tt = None
    current_temp = None
    current_lux = None
    current_param = None

    for row in ws.iter_rows(min_col=conf_min_col, min_row=conf_min_row + 1, max_row=conf_max_row):
        for cell in row:
            value = get_cell_val(ws, cell)
            # print(value, only_chars(cell.coordinate))  # Debugging
            # print(f'curr temp: {current_temp}, curr lux: {current_lux}, current param: {current_param}')  # Debugging
            if value is None:
                break
            elif isinstance(value, str) and filter_str in value.lower():
                break

            col = only_chars(cell.coordinate)
            if col == test_type_col:  # Test type
                if value is not None:
                    current_tt = value.strip().replace('\n', '')
                    try:
                        tests_seq[current_tt]
                    except KeyError:
                        tests_seq[current_tt] = {}
                    print('\n\nType: ' + current_tt)
                else:
                    current_tt = None

            if col == light_temp_col:  # Light Temperature
                if value is not None and current_tt is not None:
                    current_temp = kelvin_to_illumenant(value)
                    try:
                        tests_seq[current_tt][current_temp]
                    except KeyError:
                        tests_seq[current_tt][current_temp] = {}

                        print('value:', value)
                    print('Light Temp: ' + current_temp)
                else:
                    current_temp = None

            if col == lux_col:  # LUX
                if value is not None and current_temp is not None:
                    current_lux = only_digits(value)
                    try:
                        tests_seq[current_tt][current_temp][current_lux]
                    except KeyError:
                        tests_seq[current_tt][current_temp][current_lux] = {}
                        tests_seq[current_tt][current_temp][current_lux]['params'] = {}
                    print('- LUX: ' + str(current_lux))
                else:
                    current_lux = None

            if col == param_col:  # Params
                if value is not None and current_lux is not None:
                    current_param = value
                    tests_seq[current_tt][current_temp][current_lux]['params'][current_param] = {}
                    print('\tPARAM: ' + str(current_param))
                else:
                    current_param = None

            if col == min_col:  # Min
                if current_param is not None:
                    if current_param not in tests_seq[current_tt][current_temp][current_lux]['params']:
                        tests_seq[current_tt][current_temp][current_lux]['params'][current_param] = {}

                    tests_seq[current_tt][current_temp][current_lux]['params'][current_param]['min'] = value
                    print('\tMin: ' + str(value))

            if col == max_col:  # Max
                if current_param is not None:
                    if current_param not in tests_seq[current_tt][current_temp][current_lux]['params']:
                        tests_seq[current_tt][current_temp][current_lux]['params'][current_param] = {}

                    tests_seq[current_tt][current_temp][current_lux]['params'][current_param]['max'] = value
                    print('\tMax: ' + str(value) + '\n')

    return tests_seq


def export_to_excel_file(template_data, dest_file, add_images_bool: bool):
    # Pass template data with analysis results and requirements
    print('exporting to excel...')
    workbook = Workbook()
    sheet = workbook.active
    current_row = 1
    # Add title,, date, header etc

    current_row += 1
    current_row = xls_draw_results_table(
        template_data,
        sheet,
        1,
        current_row,
        primary_bg_color='ddebf7',
        primary_font_color='000000',
        secondary_bg_color='729fcf',
        secondary_font_color='000000',
        add_images_bool=add_images_bool
    )[1][0]

    workbook.save(filename=dest_file)


def xls_draw_results_table(template_data: dict, sheet, start_col: int, start_row: int,
                           primary_bg_color=constants.MID_COLOR.strip('#'),
                           secondary_bg_color=constants.ALTERNATE_COLOR.strip('#'),
                           primary_font_color='000000',
                           secondary_font_color=constants.BUTTON_TEXT_COLOR.strip('#'),
                           add_images_bool=True, add_border_bool=True):
    """

    :param start_col:
    :param secondary_font_color:
    :param primary_font_color:
    :param add_images_bool:
    :param template_data:
    :param sheet:
    :param start_row:
    :param primary_bg_color:
    :param secondary_bg_color:
    :param add_border_bool:
    :return: (start_col, start_row), (end_col, end_row)
    """
    center = Alignment(horizontal='center', vertical='center')

    print(template_data)

    columns = {
        'test_type': ['Test Target'],
        'image': ['Image', '', 22],
        'light_temp': ['Light'],
        'lux': ['Lux Level'],
        'param': ['Parameter'],
        'param_min': ['Min'],
        'param_max': ['Max'],
        'param_calc': ['Result'],
        'param_passfail': ['Pass/Fail'],
    }

    if not add_images_bool:
        del columns['image']

    current_row = start_row
    end_col = len(columns) - 1 + start_col

    primary_fill = PatternFill(fgColor=f"FF{primary_bg_color}", fill_type="solid")
    secondary_fill = PatternFill(fgColor=f"FF{secondary_bg_color}", fill_type="solid")

    # Styles ---

    bd = Side(style='thin', color="000000")

    header_cell_style = NamedStyle(name="header_cell")
    header_cell_style.font = Font(bold=True, color=f'FF{secondary_font_color}')
    header_cell_style.fill = secondary_fill
    header_cell_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    header_cell_style.alignment = center

    cells_style = NamedStyle(name="data_cell")
    cells_style.font = Font(bold=True, color=f'FF{primary_font_color}')
    cells_style.fill = primary_fill
    cells_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    cells_style.alignment = center

    # Styles ---

    current_col = start_col
    # Add table header
    for val in columns.values():
        try:
            val[1] = current_col
        except IndexError:
            val.append(current_col)

        cell = sheet.cell(current_row, val[1])
        cell.value = val[0]
        cell.style = header_cell_style

        data_len = len(str(val[0]))
        try:
            if val[2] < data_len:
                val[2] = data_len
        except IndexError:
            val.append(data_len)

        current_col += 1
    sheet.row_dimensions[current_row].height = 25

    current_row += 1

    print('Columns are', columns)

    test_types_list = list(template_data.keys())
    for type_num, test_type in enumerate(test_types_list):
        type_start_row = current_row
        light_colors_list = list(template_data[test_type].keys())

        for light_color_num, light_color in enumerate(light_colors_list):
            color_start_row = current_row
            luxes = list(template_data[test_type][light_color].keys())

            for lux_num, lux in enumerate(luxes):
                lux_start_row = current_row
                params = list(template_data[test_type][light_color][lux]['params'].keys())

                for param in params:
                    current_col = start_col
                    # Write row of param
                    param_templ_data = template_data[test_type][light_color][lux]['params'][param]

                    # Add test type name
                    test_type_pretty = constants.IMATEST_TEST_TYPES_FRIENDLY[test_type]
                    sheet.cell(current_row, columns['test_type'][1], test_type_pretty).style = cells_style
                    data_len = len(test_type_pretty)
                    try:
                        if columns['test_type'][2] < data_len:
                            columns['test_type'][2] = data_len
                    except IndexError:
                        columns['test_type'].append(data_len)
                    current_col += 1

                    if add_images_bool:
                        # Skip image
                        current_col += 1

                    # Add light color
                    sheet.cell(current_row, columns['light_temp'][1], light_color).style = cells_style
                    data_len = len(str(light_color))
                    try:
                        if columns['light_temp'][2] < data_len:
                            columns['light_temp'][2] = data_len + 2
                    except IndexError:
                        columns['light_temp'].append(data_len + 2)
                    current_col += 1

                    # Add lux
                    sheet.cell(current_row, columns['lux'][1], lux).style = cells_style
                    data_len = len(str(lux))
                    try:
                        if columns['lux'][2] < data_len:
                            columns['lux'][2] = data_len
                    except IndexError:
                        columns['lux'].append(data_len)
                    current_col += 1

                    # Add param
                    sheet.cell(current_row, columns['param'][1], param).style = cells_style
                    data_len = len(str(param))
                    try:
                        if columns['param'][2] < data_len:
                            columns['param'][2] = data_len + 2
                    except IndexError:
                        columns['param'].append(data_len + 2)
                    current_col += 1

                    # Add param min
                    sheet.cell(current_row, columns['param_min'][1], param_templ_data['min']).style = cells_style
                    data_len = len(str(param_templ_data['min']))
                    try:
                        if columns['param_min'][2] < data_len:
                            columns['param_min'][2] = data_len
                    except IndexError:
                        columns['param_min'].append(data_len)
                    current_col += 1

                    # Add param max
                    sheet.cell(current_row, columns['param_max'][1], param_templ_data['max']).style = cells_style
                    data_len = len(str(param_templ_data['max']))
                    try:
                        if columns['param_max'][2] < data_len:
                            columns['param_max'][2] = data_len
                    except IndexError:
                        columns['param_max'].append(data_len)
                    current_col += 1

                    # Add param result value
                    try:
                        sheet.cell(current_row, columns['param_calc'][1],
                                   format(param_templ_data['result_calculated'], '.3f')).style = cells_style
                        data_len = len(str(param_templ_data['result_calculated']))
                        try:
                            if columns['param_calc'][2] < data_len:
                                columns['param_calc'][2] = data_len
                        except IndexError:
                            columns['param_calc'].append(data_len)
                    except KeyError:
                        print(f'Missing result at {test_type}>{light_color}>{lux}>{param} -> {param_templ_data}')
                        continue
                    current_col += 1

                    # Add param Pass/Fail
                    pass_fail_cell = sheet.cell(current_row, columns['param_passfail'][1])
                    pass_fail_cell.alignment = center
                    if param_templ_data['result_pass_bool']:
                        # Passed
                        print('pass')
                        pass_fail_cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
                        pass_fail_cell.value = 'Pass'
                    else:
                        # Failed
                        print('fail')
                        pass_fail_cell.fill = PatternFill(fgColor="FF0000", fill_type="solid")
                        pass_fail_cell.value = 'Fail'
                    data_len = len(str(pass_fail_cell.value))
                    try:
                        if columns['param_passfail'][2] < data_len:
                            columns['param_passfail'][2] = data_len
                    except IndexError:
                        columns['param_passfail'].append(data_len)

                    # After each param
                    current_row += 1

                # After each lux
                if add_images_bool:
                    # Add image
                    img_cell = sheet.cell(lux_start_row, columns['image'][1])
                    img_cell.style = cells_style
                    img_file = template_data[test_type][light_color][lux]['filename']
                    print('img: ', img_file)
                    xls_import_image(img_file, sheet, img_cell)
                    # Merge image rows
                    if lux_start_row < current_row - 1:
                        sheet.merge_cells(start_column=columns['image'][1], start_row=lux_start_row,
                                          end_column=columns['image'][1], end_row=current_row - 1)

                print(f"{test_type}>{light_color}>{lux}")
                print('islast: ', (lux_num == len(luxes) - 1))
                print('merge cells would be: ')
                print(f"start_col: {columns['lux'][1]}, start_row: {lux_start_row}")
                print(f"end_col: {columns['lux'][1]}, end_row: {current_row}")
                if lux_num == len(luxes) - 1:
                    # If last one
                    pass

                if lux_start_row < current_row - 1:
                    sheet.merge_cells(start_column=columns['lux'][1], start_row=lux_start_row,
                                      end_column=columns['lux'][1], end_row=current_row - 1)

            # After each color temp
            if light_color_num == len(light_colors_list) - 1:
                # If last one
                pass

            # Merge cells
            if color_start_row < current_row - 1:
                sheet.merge_cells(start_column=columns['light_temp'][1], start_row=color_start_row,
                                  end_column=columns['light_temp'][1], end_row=current_row - 1)

        # After each Test Type
        if type_start_row < current_row - 1 :
            sheet.merge_cells(start_column=columns['test_type'][1], start_row=type_start_row,
                              end_column=columns['test_type'][1], end_row=current_row - 1)
        else:
            # current_row += 1
            continue
        if add_border_bool:
            xls_set_border(sheet, start_col, type_start_row, end_col, current_row - 1, 'thick')

        if type_num == len(test_types_list) - 1:
            # If last one
            # Merge cells
            pass
        else:
            # Add filled space
            xls_fill_cells(sheet, start_col, current_row, end_col, current_row, secondary_fill)
            current_row += 1

    print('columns after: ', columns)

    # Set columns' widths
    for col_key in columns.values():
        sheet.column_dimensions[get_column_letter(col_key[1])].width = col_key[2]

    xls_set_border(sheet, start_col, start_row, end_col, current_row - 1, 'thick')

    return (start_col, start_row), (end_col, current_row)


def xls_fill_cells(ws, start_col, start_row, end_col, end_row, fill):
    for row in ws.iter_rows(start_row, end_row, start_col, end_col):
        for cell in row:
            cell.fill = fill


def xls_set_border(ws, start_col, start_row, end_col, end_row, size='thin', color="FF000000"):
    cell_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    rows = ws[cell_range]
    side = Side(border_style=size, color=color)

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border


def xls_import_image(img_file, sheet, img_cell):
    if not os.path.isfile(img_file):
        return

    # If file is a video
    if img_file.endswith('mp4'):
        print('File ', img_file, 'is a video! Extracting frames!')
        extracted_frames = extract_video_frame(img_file, start_frame=3)
        img_file = extracted_frames[0]

    img = xls_image(img_file)
    ratio = img.width / img.height
    img.width = 150
    img.height = int(img.width / ratio)

    sheet.add_image(img, img_cell.coordinate)
