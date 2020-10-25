import json
import os
import threading

# Excel
from openpyxl import Workbook, load_workbook
from openpyxl import cell as xlcell, worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Alignment, Side
import win32com.client as win32

# Imatest
from imatest.it import ImatestLibrary, ImatestException

# Local
import src.constants as constants
from src.app.utils import kelvin_to_illumenant, only_digits, only_chars


class Report:
    def __init__(self):
        try:
            self.imatest = ImatestLibrary()
        except ImatestException.MatlabException as e:
            raise RuntimeWarning('Imatest: MathLab Error!\n', e)
        except ImatestException.LicenseException as e:
            raise RuntimeWarning('Imatest: License Error!\n', e)
        except RuntimeError as e:
            print('Imatest: Runtime Error!\n', e)

    def analyze_images(self, test_type, root_dir, images_list):
        result = None

        # root_dir = os.path.normpath(r'C:\Program Files\Imatest\v2020.2\IT\samples\python\esfriso')

        # root_dir = os.path.normpath(r'C:\Users\mms00519\Desktop\test_batch')
        # images_dir = os.path.join(root_dir, 'Images')
        # # input_file = os.path.join(images_dir, r'ESFR_50.jpg')
        # images = ['ESFR_50.jpg', 'ESFR_50.jpg']

        ini_file = os.path.join(root_dir, r'settings-p30.ini')
        op_mode = ImatestLibrary.OP_MODE_SEPARATE

        input_files = []

        for image in images_list:
            input_files.append(image)

        print('Input list: ', input_files)

        test_type_call = None
        if test_type == 'sfr':
            test_type_call = self.imatest.sfr_json
        elif test_type == 'sfrplus':
            test_type_call = self.imatest.sfrplus_json
        elif test_type == 'star':
            test_type_call = self.imatest.star_json
        elif test_type == 'colorcheck':
            test_type_call = self.imatest.colorcheck_json
        elif test_type == 'stepchart':
            test_type_call = self.imatest.stepchart_json
        elif test_type == 'wedge':
            test_type_call = self.imatest.wedge_json
        elif test_type == 'uniformity':
            test_type_call = self.imatest.uniformity_json
        elif test_type == 'distortion':
            test_type_call = self.imatest.distortion_json
        elif test_type == 'esfriso':
            test_type_call = self.imatest.esfriso_json
        elif test_type == 'blemish':
            test_type_call = self.imatest.blemish_json
        elif test_type == 'dotpattern':
            test_type_call = self.imatest.dotpattern_json
        elif test_type == 'color_tone':
            test_type_call = self.imatest.color_tone_json
        elif test_type == 'checkerboard':
            test_type_call = self.imatest.checkerboard_json
        elif test_type == 'random':
            test_type_call = self.imatest.random_json
        elif test_type == 'sfrreg':
            test_type_call = self.imatest.sfrreg_json
        elif test_type == 'logfc':
            test_type_call = self.imatest.logfc_json
        # Only other modules are "Arbitrary Charts" and "OIS" - they want different arguments

        # Call to esfriso using Op Mode Standard, with ini file argument and JSON output
        # with multiple files.
        # input_file argument is a list containing the full paths to several images.
        # Output is a JSON string containing the result of the last image

        try:
            result = test_type_call(input_file=input_files,
                                    root_dir=root_dir,
                                    op_mode=op_mode,
                                    ini_file=ini_file)

            print(result)
        except ImatestException as iex:
            if iex.error_id == ImatestException.FloatingLicenseException:
                print("All floating license seats are in use.  Exit Imatest on another computer and try again.")
            elif iex.error_id == ImatestException.LicenseException:
                print("License Exception: " + iex.message)
            else:
                print(iex.message)
        except Exception as ex:
            print(str(ex))

        return json.loads(result)  # Return readable json object

    def analyze_images_parallel(self, images, ini_file, num_processes: int = 4):
        tasks = []

        for device_serial in images.keys():
            for test_num, test_dict in enumerate(images[device_serial]):
                print(f"ADDING TASK - {test_dict['analysis_type']}\n{test_dict['image_files']}")
                tasks.append(
                    self.imatest.new_parallel_task(
                        image_files=test_dict['image_files'],
                        analysis_type=self.imatest_analysis_type(test_dict['analysis_type'])
                    )
                )

        result = self.imatest.parallel_analyzer(tasks=tasks, ini_file=ini_file, run_parallel=True,
                                                num_workers=num_processes)

        return json.loads(result)  # Return readable json object

    def imatest_analysis_type(self, test_type):
        # Filter out special characters and spaces
        test_type = ''.join(filter(str.isalnum, test_type.lower()))

        for tt in constants.IMATEST_TEST_TYPES.keys():
            if test_type == tt:
                print(f'For analysis type {test_type} we found {getattr(self.imatest, constants.IMATEST_TEST_TYPES[tt])}')  # DEBUG PRINT
                return getattr(self.imatest, constants.IMATEST_TEST_TYPES[tt])
        # if it still hasn't returned, then will come through here
        print('Analysis Type not found in IMATEST_TEST_TYPES: ', test_type)

    def __del__(self):
        print("Terminating Imatest Library")
        # When finished terminate the library
        self.imatest.terminate_library()

    # Static Methods
    @staticmethod
    def recurse_dict(d, keys=()):
        if type(d) == dict:
            for k in d:
                for rv in Report.recurse_dict(d[k], keys + (k,)):
                    yield rv
        else:
            yield (keys, d)

    @staticmethod
    def update_imatest_params():
        report_obj = Report()
        images_dict = {'test_serial': []}
        tests_params = {}
        curr_progress = 0
        ini_file = os.path.join(constants.ROOT_DIR, 'images', 'imatest', 'ini_file', 'imatest-v2.ini')

        prog_step = 100/len(constants.IMATEST_TEST_TYPES.keys())/3

        # Prepare list for passing to image analyzer
        tests_list = images_dict['test_serial']
        for test_name in constants.IMATEST_TEST_TYPES.keys():
            img_file = os.path.join(constants.ROOT_DIR, 'images', 'imatest', f'{test_name}_example')
            if os.path.exists(img_file + '.jpg'):
                img_file += '.jpg'
            elif os.path.exists(img_file + '.png'):
                img_file += '.png'
            else:
                print('FAILED TO FIND IMAGE FOR', img_file)
                continue

            new_dict = {
                'analysis_type': test_name,
                'image_files': [img_file]
            }

            tests_list.append(new_dict)

        print('Created images dict:\n', images_dict)
        result = report_obj.analyze_images_parallel(images_dict, ini_file)
        # open output file for writing
        result_out_file = os.path.join(constants.DATA_DIR, 'imatest_all_tests_results.json')
        with open(result_out_file, 'w') as outfile:
            json.dump(result, outfile)

        print('Analysis Result:\n', result)

        # Parse received list to params file
        filter_params = [
            'dateRun', 'ini_time_size', 'version', 'buildDate', 'title',
            'image_file', 'image_path_name', 'build', 'EXIF_results',
            'errorID', 'errorMessage', 'errorReport',
            '_ArrayType_', '_ArraySize_', '_ArrayData_'
        ]

        for res_dict in result:
        #for res_dict in images_analysis_readable:
            current_type = None
            for key, value in Report.recurse_dict(res_dict['data']):
                if current_type is None:
                    # First find the title
                    # Turns out dotpattern has neither title, nor image_file keys..
                    # if key[0] == 'title' or key[0] == 'image_file':
                    #     current_type = value.split('_')[0]
                    if key[0] == 'version':
                        current_type = value.split('  ')[-1].split('_')[0].replace(' ', '').lower()
                else:
                    param_name = '>'.join(key)
                    val_type = type(value).__name__
                    if val_type == type(str).__name__ or key[-1] in filter_params:
                        # Skip string or filtered params
                        print(f'Skipping {key[-1]} because it is a string or in filter!')
                        continue
                    try:
                        tests_params[current_type]
                    except KeyError:
                        tests_params[current_type] = {}
                        tests_params[current_type][param_name] = val_type
                    else:
                        tests_params[current_type][param_name] = val_type

        # open output file for writing
        params_out_file = os.path.join(constants.DATA_DIR, 'imatest_params.json')
        with open(params_out_file, 'w') as outfile:
            json.dump(tests_params, outfile)

    @staticmethod
    def update_imatest_params_threaded():
        threading.Thread(target=Report.update_imatest_params,
                         args=(),
                         daemon=True).start()

    @staticmethod
    def xls_to_xlsx(xls_file) -> str:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(xls_file)
        new_file = xls_file + "x"

        wb.SaveAs(new_file, FileFormat=51)  # FileFormat = 51 is for .xlsx extension
        wb.Close()  # FileFormat = 56 is for .xls extension
        excel.Application.Quit()

        return new_file

    @staticmethod
    def within_range(bounds: tuple, cell: xlcell) -> bool:
        column_start, row_start, column_end, row_end = bounds
        row = cell.row
        if row_start <= row <= row_end:
            column = cell.column
            if column_start <= column <= column_end:
                return True
        return False

    @staticmethod
    def get_value_merged(sheet: worksheet, cell: xlcell) -> any:
        for merged in sheet.merged_cells:
            if Report.within_range(merged.bounds, cell):
                return sheet.cell(merged.min_row, merged.min_col).value
        return cell.value

    @staticmethod
    def is_merged(cell) -> bool:
        if type(cell).__name__ == 'MergedCell':
            return True
        else:
            return False

    @staticmethod
    def get_cell_val(ws, cell):
        return Report.get_value_merged(ws, cell) if Report.is_merged(cell) else cell.value

    @staticmethod
    def parse_excel_template(excel_file) -> dict:
        # Config
        header_mapping = {
            'test_type': "test target",
            'light_temp': "light",
            'lux': "lux",
            'param': "param",
            'min': "min",
            'max': "max"
        }
        conf_min_col = 1
        conf_min_row = 4
        conf_max_row = 77
        filter_str = 'functional'
        # Config end

        excel_file = os.path.normpath(excel_file)
        print(f'Parsing "{excel_file}"...')

        ext = excel_file.split('.')[-1]

        if ext == 'xls':
            print("Got .XLS.. Converting it to XLSX")
            excel_file = Report.xls_to_xlsx(excel_file)

        wb = load_workbook(filename=excel_file)
        ws = wb.worksheets[0]

        tests_seq = {}

        header = list(ws.rows)[1]

        test_type_col = None
        light_temp_col = None
        lux_col = None
        param_col = None
        min_col = None
        max_col = None

        for cell in header:
            if header_mapping['test_type'] in cell.value.lower():
                test_type_col = cell.column_letter

            if header_mapping['light_temp'] in cell.value.lower():
                light_temp_col = cell.column_letter

            if header_mapping['lux'] in cell.value.lower():
                lux_col = cell.column_letter

            if header_mapping['param'] in cell.value.lower():
                param_col = cell.column_letter

            if header_mapping['min'] in cell.value.lower():
                min_col = cell.column_letter

            if header_mapping['max'] in cell.value.lower():
                max_col = cell.column_letter

        print(
            f"Test Type: {test_type_col}, \nLight Temp: {light_temp_col}, \nLux: {lux_col}, \nParams: {param_col}, \nMin: {min_col}, \nMax: {max_col}\n")

        current_tt = None
        current_temp = None
        current_lux = None
        current_param = None

        for row in ws.iter_rows(min_col=conf_min_col, min_row=conf_min_row, max_row=conf_max_row):
            for cell in row:
                value = Report.get_cell_val(ws, cell)
                # print(value, only_chars(cell.coordinate))  # Debugging
                # print(f'curr temp: {current_temp}, curr lux: {current_lux}, current param: {current_param}')  # Debugging
                if value is not None and isinstance(value, str) and filter_str in value.lower():
                    break

                col = only_chars(cell.coordinate)
                if col == test_type_col:  # Test type
                    if value is not None:
                        current_tt = value.strip().replace('\n', '')
                        try:
                            tests_seq[current_tt]
                        except KeyError:
                            tests_seq[current_tt] = {}
                        print('\n\nType: ' + current_tt)
                    else:
                        current_tt = None

                if col == light_temp_col:  # Light Temperature
                    if value is not None and current_tt is not None:
                        current_temp = kelvin_to_illumenant(value)
                        try:
                            tests_seq[current_tt][current_temp]
                        except KeyError:
                            tests_seq[current_tt][current_temp] = {}
                        print('Light Temp: ' + current_temp)
                    else:
                        current_temp = None

                if col == lux_col:  # LUX
                    if value is not None and current_temp is not None:
                        current_lux = only_digits(value)
                        try:
                            tests_seq[current_tt][current_temp][current_lux]
                        except KeyError:
                            tests_seq[current_tt][current_temp][current_lux] = {}
                            tests_seq[current_tt][current_temp][current_lux]['params'] = {}
                        print('- LUX: ' + str(current_lux))
                    else:
                        current_lux = None

                if col == param_col:  # Params
                    if value is not None and current_lux is not None:
                        current_param = value
                        tests_seq[current_tt][current_temp][current_lux]['params'][current_param] = {}
                        print('\tPARAM: ' + str(current_param))
                    else:
                        current_param = None

                if col == min_col:  # Min
                    if current_param is not None:
                        if current_param not in tests_seq[current_tt][current_temp][current_lux]['params']:
                            tests_seq[current_tt][current_temp][current_lux]['params'][current_param] = {}

                        tests_seq[current_tt][current_temp][current_lux]['params'][current_param]['min'] = value
                        print('\tMin: ' + str(value))

                if col == max_col:  # Max
                    if current_param is not None:
                        if current_param not in tests_seq[current_tt][current_temp][current_lux]['params']:
                            tests_seq[current_tt][current_temp][current_lux]['params'][current_param] = {}

                        tests_seq[current_tt][current_temp][current_lux]['params'][current_param]['max'] = value
                        print('\tMax: ' + str(value) + '\n')

        return tests_seq

    @staticmethod
    def generate_lights_seqs(req_dict):
        lights_list = []

        for test_type in req_dict.keys():
            lights_list.append(
                {
                    'test_type': test_type,
                    'lights_seq': {}
                }
            )
            for light_type in req_dict[test_type].keys():
                lights_list[len(lights_list) - 1]['lights_seq'][light_type] = []
                for lux in req_dict[test_type][light_type].keys():
                    lights_list[len(lights_list) - 1]['lights_seq'][light_type].append(lux)

        return lights_list

    @staticmethod
    def get_list_average(list_in: list):
        if not isinstance(list_in, list):
            return

        result = 0
        divider = 0

        for item in list_in:
            if isinstance(item, float) or isinstance(item, int) or (isinstance(item, str) and item.isdigit()):
                if isinstance(item, str) and item.isdigit():
                    result += int(item)
                    divider += 1
                    continue
                result += item
                divider += 1

        return result/divider

    @staticmethod
    def analyze_images_test_results(template_data):
        for test_type in template_data.keys():
            for light_temp in template_data[test_type].keys():
                for lux in template_data[test_type][light_temp].keys():
                    img_file_path = template_data[test_type][light_temp][lux]['filename']
                    img_file_name = os.path.basename(img_file_path)
                    img_results_path = os.path.join(os.path.dirname(img_file_path), 'Results')
                    img_json_filename = [f for f in os.listdir(img_results_path) if
                                         f.startswith(img_file_name.split('.')[0]) and f.endswith('.json')]

                    if len(img_json_filename) < 1:
                        continue

                    img_json_file = os.path.join(img_results_path, img_json_filename[0])

                    print('img jsonfile: ', img_json_file)

                    print('Now at img: ', img_file_name)

                    with open(img_json_file) as json_file:
                        image_analysis_readable = json.load(json_file)
                    image_analysis_readable = image_analysis_readable[list(image_analysis_readable.keys())[0]]

                    for param in template_data[test_type][light_temp][lux]['params'].keys():
                        params_depth = param.split('>')
                        for num, param_piece in enumerate(params_depth):
                            if num == 0:
                                # At beginning set equal to parent of param (possibly)
                                param_val = image_analysis_readable[param_piece]
                            elif num != len(params_depth) - 1:
                                # If not at end yet and not at beginning
                                param_val = param_val[param_piece]

                            if num == len(params_depth) - 1:
                                # If last one -> return
                                # Full name of param: param,
                                # last part of param: param_piece,
                                # param value: param_val
                                curr_param_dict = template_data[test_type][light_temp][lux]["params"][param]
                                param_val_calc = Report.get_list_average(param_val)
                                curr_param_dict['result'] = param_val
                                curr_param_dict['result_calculated'] = param_val_calc

                                print(f'param {param} is: ', param_val)
                                print(f'calculated value is: {param_val_calc}')
                                print(f'min is: {curr_param_dict["min"]}')
                                print(f'max is: {curr_param_dict["max"]}')
                                if param_val_calc > curr_param_dict["min"] and param_val_calc < curr_param_dict["max"]:
                                    curr_param_dict['result_pass_bool'] = True
                                    print('PASS!\n')
                                else:
                                    curr_param_dict['result_pass_bool'] = False
                                    print('FAIL!\n')

        return template_data

    @staticmethod
    def export_to_excel_file(template_data, dest_file):
        # Pass template data with analysis results and requirements
        print('exporting to excel...')
        workbook = Workbook()
        sheet = workbook.active
        current_row = 1
        # Add title,, date, header etc


        current_row += 1
        current_row = Report.xls_draw_results_table(
            template_data,
            sheet,
            2,
            current_row,
            primary_bg_color=constants.MID_COLOR.strip('#'),
            secondary_bg_color=constants.ALTERNATE_COLOR.strip('#'),
            font_color=constants.TEXT_COLOR.strip('#')
        )[1][0]

        workbook.save(filename=dest_file)

    @staticmethod
    def xls_draw_results_table(template_data, sheet, start_col, start_row,
                               primary_bg_color=constants.MID_COLOR, secondary_bg_color=constants.ALTERNATE_COLOR,
                               font_color=constants.TEXT_COLOR,
                               add_border_bool=True):
        """

        :param template_data:
        :param sheet:
        :param start_row:
        :param primary_bg_color:
        :param secondary_bg_color:
        :param font_color:
        :param add_border_bool:
        :return: (start_col, start_row), (end_col, end_row)
        """
        center = Alignment(horizontal='center', vertical='center')

        print(template_data)

        columns = {
            'test_type': ['Test Target'],
            'image': ['Image', '', 20],
            'light_temp': ['Light'],
            'lux': ['Lux Level'],
            'param': ['Parameter'],
            'param_min': ['Min'],
            'param_max': ['Max'],
            'param_calc': ['Result'],
            'param_passfail': ['Pass/Fail'],
        }

        current_row = start_row
        end_col = len(columns)-1 + start_col

        columns_count = 0
        current_col = start_col
        # Add table header
        for val in columns.values():
            try:
                val[1] = current_col
            except IndexError:
                val.append(current_col)

            cell = sheet.cell(current_row, val[1])
            cell.value = val[0]
            cell.font = Font(bold=True)
            cell.alignment = center

            data_len = len(str(val[0]))
            try:
                if val[2] < data_len:
                    val[2] = data_len
            except IndexError:
                val.append(data_len)

            current_col += 1

        current_row += 1

        print('Columns are', columns)

        secondary_fill = PatternFill(fgColor=f"FF{secondary_bg_color}", fill_type="solid")

        test_types_list = list(template_data.keys())
        for type_num, test_type in enumerate(test_types_list):
            type_start_row = current_row
            light_colors_list = list(template_data[test_type].keys())
            for light_color_num, light_color in enumerate(light_colors_list):
                color_start_row = current_row
                luxes = list(template_data[test_type][light_color].keys())
                for lux_num, lux in enumerate(luxes):
                    lux_start_row = current_row
                    params = list(template_data[test_type][light_color][lux]['params'].keys())
                    for param in params:
                        current_col = start_col
                        # Write row of param
                        param_templ_data = template_data[test_type][light_color][lux]['params'][param]

                        # Add test type name
                        sheet.cell(current_row, columns['test_type'][1], test_type).alignment = center
                        data_len = len(str(test_type))
                        try:
                            if columns['test_type'][2] < data_len:
                                columns['test_type'][2] = data_len
                        except IndexError:
                            columns['test_type'].append(data_len)
                        current_col += 1

                        # Add image
                        sheet.cell(current_row, columns['image'][1], 'image').alignment = center
                        current_col += 1

                        # Add light color
                        sheet.cell(current_row, columns['light_temp'][1], light_color).alignment = center
                        data_len = len(str(light_color))
                        try:
                            if columns['light_temp'][2] < data_len:
                                columns['light_temp'][2] = data_len
                        except IndexError:
                            columns['light_temp'].append(data_len)
                        current_col += 1

                        # Add lux
                        sheet.cell(current_row, columns['lux'][1], lux).alignment = center
                        data_len = len(str(lux))
                        try:
                            if columns['lux'][2] < data_len:
                                columns['lux'][2] = data_len
                        except IndexError:
                            columns['lux'].append(data_len)
                        current_col += 1

                        # Add param
                        sheet.cell(current_row, columns['param'][1], param).alignment = center
                        data_len = len(str(param))
                        try:
                            if columns['param'][2] < data_len:
                                columns['param'][2] = data_len
                        except IndexError:
                            columns['param'].append(data_len)
                        current_col += 1

                        # Add param min
                        sheet.cell(current_row, columns['param_min'][1], param_templ_data['min']).alignment = center
                        data_len = len(str(param_templ_data['min']))
                        try:
                            if columns['param_min'][2] < data_len:
                                columns['param_min'][2] = data_len
                        except IndexError:
                            columns['param_min'].append(data_len)
                        current_col += 1

                        # Add param max
                        sheet.cell(current_row, columns['param_max'][1], param_templ_data['max']).alignment = center
                        data_len = len(str(param_templ_data['max']))
                        try:
                            if columns['param_max'][2] < data_len:
                                columns['param_max'][2] = data_len
                        except IndexError:
                            columns['param_max'].append(data_len)
                        current_col += 1
                        current_col += 1

                        # Add param result value
                        try:
                            sheet.cell(current_row, columns['param_calc'][1],
                                       param_templ_data['result_calculated']).alignment = center
                            data_len = len(str(param_templ_data['result_calculated']))
                            try:
                                if columns['param_calc'][2] < data_len:
                                    columns['param_calc'][2] = data_len
                            except IndexError:
                                columns['param_calc'].append(data_len)
                        except KeyError:
                            print(f'Missing result at {test_type}>{light_color}>{lux}>{param} -> {param_templ_data}')
                            continue
                        current_col += 1

                        # Add param Pass/Fail
                        pass_fail_cell = sheet.cell(current_row, columns['param_passfail'][1])
                        pass_fail_cell.alignment = center
                        if param_templ_data['result_pass_bool']:
                            # Passed
                            print('pass')
                            pass_fail_cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
                            pass_fail_cell.value = 'Pass'
                        else:
                            # Failed
                            print('fail')
                            pass_fail_cell.fill = PatternFill(fgColor="FF0000", fill_type="solid")
                            pass_fail_cell.value = 'Fail'
                        data_len = len(str(pass_fail_cell.value))
                        try:
                            if columns['param_passfail'][2] < data_len:
                                columns['param_passfail'][2] = data_len
                        except IndexError:
                            columns['param_passfail'].append(data_len)

                        # After each param
                        current_row += 1
                        print(f'\t\t\tparam done (col: {current_col}, row: {current_row}): ', param)

                        # Merge cells
                        # sheet.merge_cells(start_column=columns['lux'][1], start_row=lux_start_row,
                        #                   end_column=columns['lux'][1], end_row=current_row - 1)
                        # sheet.merge_cells(start_column=columns['light_temp'][1], start_row=color_start_row,
                        #                   end_column=columns['light_temp'][1], end_row=current_row - 1)
                        # sheet.merge_cells(start_column=columns['test_type'][1], start_row=type_start_row,
                        #                   end_column=columns['test_type'][1], end_row=current_row - 1)


                    # After each lux
                    if lux_num == len(luxes)-1:
                        # If last one
                        pass
                    print('\t\tlux done: ', lux)
                    # current_row += 1
                # After each color temp
                if light_color_num == len(light_colors_list)-1:
                    # If last one
                    pass

                print('\tcolor_temp done: ', light_color)
                # current_row += 1
            # After each Test Type
            if type_num == len(test_types_list) - 1:
                # If last one
                pass
            else:
                Report.xls_set_border(sheet, start_col, type_start_row, current_row-1, end_col)
                Report.xls_fill_cells(sheet, start_col, current_row, end_col, current_row, secondary_fill)
                current_row += 1

            print('test_type done: ', test_type)

        print('columns after: ', columns)

        # Set columns' widths
        for col_key in columns.values():
            sheet.column_dimensions[get_column_letter(col_key[1])].width = col_key[2]
            # sheet.column_dimensions[get_column_letter(col_key[1])].bestFit = True
            # sheet.column_dimensions[get_column_letter(col_key[1])].auto_size = True

        Report.xls_set_border(sheet, start_col, start_row, end_col, current_row, 'thick')

        return (start_col, start_row), (end_col, current_row)

    @staticmethod
    def xls_fill_cells(ws, start_col, start_row, end_col, end_row, fill):
        for row in ws.iter_rows(start_row, end_row, start_col, end_col):
            for cell in row:
                cell.fill = fill


    @staticmethod
    def xls_set_border(ws, start_col, start_row, end_col, end_row, size='thin', color="FF000000"):
        cell_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        rows = ws[cell_range]
        side = Side(border_style=size, color=color)

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border